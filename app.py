import io
import re
import calendar
import datetime as dt
from dataclasses import dataclass
from typing import Optional, Set, Dict, List, Tuple

import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import holidays
from dateutil import parser as date_parser


# ----------------------------
# Helpers: parsing and counting
# ----------------------------

PT_DOW = {
    "SEG": 0,
    "TER": 1,
    "QUA": 2,
    "QUI": 3,
    "SEX": 4,
    "SAB": 5,
    "SÁB": 5,
    "DOM": 6,
}

DOW_TOKENS_RE = re.compile(r"(SEG|TER|QUA|QUI|SEX|SAB|SÁB|DOM)", re.IGNORECASE)

# Regex to detect month-year from column headers like "DIAS ÚTEIS 01.2026 (ESCALA ANTIGA)"
MONTHYEAR_RE = re.compile(r"(\d{1,2})\.(\d{4})")


@dataclass
class RowResult:
    ok: bool
    error: Optional[str] = None


def safe_parse_date(value) -> Optional[dt.date]:
    """Parse Excel cell that may be date/datetime or string."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    try:
        return date_parser.parse(s, dayfirst=True).date()
    except Exception:
        return None


def parse_schedule_days(text: str) -> Optional[Set[int]]:
    """
    Convert schedule description to set of working weekdays (0=SEG ... 6=DOM).
    Supported:
      - "SEG A SEX", "SEG A SAB", "TER A DOM" (wrap supported)
      - list of days: "SEG TER QUA"
      - "FOLGA TER e DOM" => work = all - {TER, DOM}
    """
    if text is None:
        return None
    s = str(text).upper()

    tokens = DOW_TOKENS_RE.findall(s)
    if not tokens:
        return None

    tok = [t.upper() for t in tokens]
    tok = ["SAB" if t == "SÁB" else t for t in tok]

    if "FOLGA" in s:
        off = {PT_DOW[t] for t in tok if t in PT_DOW}
        return set(range(7)) - off

    if " A " in s and len(tok) >= 2:
        a = PT_DOW.get(tok[0])
        b = PT_DOW.get(tok[1])
        if a is None or b is None:
            return None
        if a <= b:
            return set(range(a, b + 1))
        return set(list(range(a, 7)) + list(range(0, b + 1)))

    days = {PT_DOW[t] for t in tok if t in PT_DOW}
    return days if days else None


def month_bounds(year: int, month: int) -> Tuple[dt.date, dt.date]:
    last = calendar.monthrange(year, month)[1]
    return dt.date(year, month, 1), dt.date(year, month, last)


def month_start_from_day(year: int, month: int, start_day: int) -> dt.date:
    """
    Use ONLY the day number from 'INÍCIO ESCALA NOVA' as the start day for each month/year.
    If start_day doesn't exist in that month, returns a date after month end (so count becomes 0).
    """
    last = calendar.monthrange(year, month)[1]
    if start_day > last:
        return dt.date(year, month, last) + dt.timedelta(days=1)
    return dt.date(year, month, start_day)


def count_workdays(
    start: dt.date,
    end: dt.date,
    working_days: Set[int],
    holiday_set: Set[dt.date],
) -> int:
    if start > end:
        return 0
    cnt = 0
    d = start
    one = dt.timedelta(days=1)
    while d <= end:
        if (d.weekday() in working_days) and (d not in holiday_set):
            cnt += 1
        d += one
    return cnt


def find_header_row_and_map(ws: Worksheet) -> Tuple[int, Dict[str, int]]:
    """
    Find header row by scanning first ~50 rows and mapping header text -> column index.
    Header_map keys are normalized (upper, strip).
    """
    for r in range(1, 51):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not values:
            continue
        normalized = []
        for v in values:
            if v is None:
                normalized.append("")
            else:
                normalized.append(str(v).strip().upper())

        if "INÍCIO ESCALA NOVA" in normalized and "ESCALA NOVA" in normalized:
            m: Dict[str, int] = {}
            for idx, name in enumerate(normalized, start=1):
                if name:
                    m[name] = idx
            return r, m

    raise ValueError("Não encontrei a linha de cabeçalho (preciso de 'INÍCIO ESCALA NOVA' e 'ESCALA NOVA').")


def get_column_if_exists(header_map: Dict[str, int], header_name: str) -> Optional[int]:
    """Return column index if exists; do NOT create new columns."""
    return header_map.get(header_name.strip().upper())


def build_brazil_holidays(years: Set[int]) -> Set[dt.date]:
    """National Brazil holidays for given years using python-holidays."""
    if not years:
        return set()
    br = holidays.Brazil(years=years)
    return {d for d in br.keys()}


def parse_extra_holidays(text: str) -> Set[dt.date]:
    """Parse user-entered holidays (1 per line)."""
    out: Set[dt.date] = set()
    if not text:
        return out
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        try:
            out.add(date_parser.parse(s, dayfirst=True).date())
        except Exception:
            pass
    return out


def extract_month_year_from_header(header: str) -> Optional[Tuple[int, int]]:
    """
    Extract (year, month) from something that contains 'MM.AAAA'.
    Returns (year, month).
    """
    if not header:
        return None
    m = MONTHYEAR_RE.search(str(header))
    if not m:
        return None
    month = int(m.group(1))
    year = int(m.group(2))
    if 1 <= month <= 12:
        return year, month
    return None


def detect_months_from_existing_output_columns(header_map: Dict[str, int]) -> List[Tuple[int, int]]:
    """
    Detect which months exist in the Excel by looking at existing output columns:
      - DIAS ÚTEIS MM.AAAA (ESCALA ANTIGA)
      - DIAS ÚTEIS MM.AAAA (ESCALA NOVA)
      - DIAS DEVIDOS MM.AAAA
    We will calculate ONLY for months that already appear in at least one of these columns.
    """
    months: Set[Tuple[int, int]] = set()
    for h in header_map.keys():
        # Only look at columns that match outputs or the old scale column pattern
        if ("DIAS ÚTEIS" in h) or ("DIAS DEVIDOS" in h):
            ym = extract_month_year_from_header(h)
            if ym:
                months.add(ym)
    return sorted(months, key=lambda x: (x[0], x[1]))


def find_old_scale_column_for_month(header_map: Dict[str, int], year: int, month: int) -> Optional[int]:
    """
    Old scale column is expected to be: 'ESCALA MM.AAAA'
    """
    old_scale_header = f"ESCALA {month:02d}.{year}".upper()
    return header_map.get(old_scale_header)


# ----------------------------
# Main processing
# ----------------------------

def process_workbook(
    file_bytes: bytes,
    sheet_names: Optional[List[str]],
    extra_holidays: Set[dt.date],
) -> Tuple[bytes, List[str]]:
    """
    Load xlsx from bytes, update selected sheets, return updated file bytes and logs.

    Key changes requested:
    - Use ONLY the DAY from 'INÍCIO ESCALA NOVA' for start day in each month (ignore its month/year).
    - Calculate ONLY for columns that already exist in the Excel (do not create new columns).
    """
    logs: List[str] = []
    bio = io.BytesIO(file_bytes)
    wb = load_workbook(bio)
    target_sheets = sheet_names or wb.sheetnames

    updated_any = False

    for sname in target_sheets:
        if sname not in wb.sheetnames:
            continue

        ws = wb[sname]

        try:
            header_row, header_map = find_header_row_and_map(ws)
        except Exception as e:
            logs.append(f"[ERRO] Aba '{sname}': {e}")
            continue

        # Detect which months we should calculate based on EXISTING output columns
        months_to_calc = detect_months_from_existing_output_columns(header_map)
        if not months_to_calc:
            logs.append(f"[AVISO] Aba '{sname}': não encontrei colunas de saída (DIAS ÚTEIS/DIAS DEVIDOS) com MM.AAAA. Nada a calcular.")
            continue

        # We need ESCALA NOVA and INÍCIO ESCALA NOVA
        if "ESCALA NOVA" not in header_map or "INÍCIO ESCALA NOVA" not in header_map:
            logs.append(f"[ERRO] Aba '{sname}': faltam colunas obrigatórias 'ESCALA NOVA' e/ou 'INÍCIO ESCALA NOVA'.")
            continue

        col_new = header_map["ESCALA NOVA"]
        col_start = header_map["INÍCIO ESCALA NOVA"]

        # Collect years for holidays for this sheet
        years_needed = {y for (y, m) in months_to_calc}
        years_needed |= {d.year for d in extra_holidays}
        holiday_set = build_brazil_holidays(years_needed) | set(extra_holidays)

        # Pre-map output columns that exist (so we don't create anything)
        out_cols: Dict[Tuple[int, int], Dict[str, Optional[int]]] = {}
        missing_cols_msgs: List[str] = []

        for (yy, mm) in months_to_calc:
            h_old = f"DIAS ÚTEIS {mm:02d}.{yy} (ESCALA ANTIGA)"
            h_new = f"DIAS ÚTEIS {mm:02d}.{yy} (ESCALA NOVA)"
            h_due = f"DIAS DEVIDOS {mm:02d}.{yy}"

            c_old = get_column_if_exists(header_map, h_old)
            c_new = get_column_if_exists(header_map, h_new)
            c_due = get_column_if_exists(header_map, h_due)

            out_cols[(yy, mm)] = {"old": c_old, "new": c_new, "due": c_due}

            # log missing (not created)
            miss = []
            if c_old is None: miss.append(h_old)
            if c_new is None: miss.append(h_new)
            if c_due is None: miss.append(h_due)
            if miss:
                missing_cols_msgs.append(f"{mm:02d}.{yy}: " + " | ".join(miss))

        # TOTAL DIAS DEVIDOS (optional) - only write if exists
        c_total = get_column_if_exists(header_map, "TOTAL DIAS DEVIDOS")

        if missing_cols_msgs:
            logs.append(f"[INFO] Aba '{sname}': algumas colunas de saída não existem e NÃO serão criadas. ({'; '.join(missing_cols_msgs)})")

        processed = 0
        errors = 0

        last_row = ws.max_row
        for r in range(header_row + 1, last_row + 1):
            v_new = ws.cell(row=r, column=col_new).value
            v_start = ws.cell(row=r, column=col_start).value

            # Skip fully empty lines
            if v_new is None and v_start is None:
                continue

            start_date = safe_parse_date(v_start)
            if start_date is None:
                errors += 1
                continue

            start_day = start_date.day  # <-- ONLY DAY is used

            days_new = parse_schedule_days(v_new)
            if days_new is None:
                errors += 1
                continue

            total_due = 0

            for (yy, mm) in months_to_calc:
                # Find old scale column for THIS month (ESCALA MM.AAAA) — must exist to compute old vs new
                col_old = find_old_scale_column_for_month(header_map, yy, mm)
                if col_old is None:
                    # if there's no old scale column for this month, we cannot compare; skip this month
                    continue

                v_old = ws.cell(row=r, column=col_old).value
                days_old = parse_schedule_days(v_old)
                if days_old is None:
                    # can't compute for this month
                    continue

                m_start, m_end = month_bounds(yy, mm)
                calc_start = month_start_from_day(yy, mm, start_day)  # <-- day inside month/year from headers

                old_cnt = count_workdays(calc_start, m_end, days_old, holiday_set)
                new_cnt = count_workdays(calc_start, m_end, days_new, holiday_set)
                due = old_cnt - new_cnt

                total_due += due

                # write only if those columns exist
                cols = out_cols.get((yy, mm), {})
                if cols.get("old") is not None:
                    ws.cell(row=r, column=cols["old"]).value = old_cnt
                if cols.get("new") is not None:
                    ws.cell(row=r, column=cols["new"]).value = new_cnt
                if cols.get("due") is not None:
                    ws.cell(row=r, column=cols["due"]).value = due

            if c_total is not None:
                ws.cell(row=r, column=c_total).value = total_due

            processed += 1

        logs.append(f"[OK] Aba '{sname}': {processed} linhas processadas, {errors} linhas com erro (data/escala inválida).")
        updated_any = True

    if not updated_any:
        logs.append("[AVISO] Nenhuma aba foi atualizada. Verifique se os cabeçalhos existem e se há colunas de saída (DIAS ÚTEIS/DIAS DEVIDOS) com MM.AAAA.")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), logs


# ----------------------------
# Streamlit UI
# ----------------------------

st.set_page_config(page_title="Atualizador de Escalas (Excel)", layout="centered")

st.title("Atualizar planilha de alteração de escala")
st.caption("Faça upload do .xlsx, clique em **ATUALIZAR PLANILHA** e baixe o arquivo atualizado (mesma formatação, só valores).")

uploaded = st.file_uploader("Envie sua planilha (.xlsx)", type=["xlsx"])

extra_holidays_text = st.text_area(
    "Feriados extras (opcional) — 1 por linha (ex: 25/01/2026)",
    value="",
    placeholder="Ex:\n25/01/2026\n16/07/2026",
    height=120,
)

process_all_sheets = st.checkbox("Processar TODAS as abas (recomendado)", value=True)

selected_sheets = None
if uploaded is not None:
    try:
        wb_preview = load_workbook(io.BytesIO(uploaded.getvalue()), read_only=True)
        sheetnames = wb_preview.sheetnames
        wb_preview.close()
    except Exception as e:
        st.error(f"Não consegui abrir esse arquivo: {e}")
        st.stop()

    if not process_all_sheets:
        selected_sheets = st.multiselect(
            "Selecione as abas para processar",
            options=sheetnames,
            default=sheetnames[:1],
        )

btn = st.button("ATUALIZAR PLANILHA", type="primary", disabled=(uploaded is None))

if btn and uploaded is not None:
    with st.spinner("Processando..."):
        extras = parse_extra_holidays(extra_holidays_text)
        updated_bytes, logs = process_workbook(
            file_bytes=uploaded.getvalue(),
            sheet_names=(None if process_all_sheets else selected_sheets),
            extra_holidays=extras,
        )

    st.subheader("Log")
    for line in logs:
        if line.startswith("[ERRO]"):
            st.error(line)
        elif line.startswith("[AVISO]"):
            st.warning(line)
        elif line.startswith("[INFO]"):
            st.info(line)
        else:
            st.success(line)

    filename = uploaded.name
    st.download_button(
        "Baixar planilha atualizada",
        data=updated_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.divider()
st.markdown(
    """
**Como funciona (atualizado)**
- Usa **somente o DIA** do campo **INÍCIO ESCALA NOVA** (ignora mês/ano do início).
- O mês/ano vem das colunas existentes (ex.: `DIAS ÚTEIS 02.2026 ...`).
- **Não cria colunas novas**: só preenche as que já existem no seu Excel.
- Desconta feriados nacionais do Brasil automaticamente (biblioteca `holidays`) + feriados extras opcionais.
"""
)
